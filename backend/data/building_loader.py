"""Building Loader — TL_SGCO_RNADR_MST + 총괄표제부 xlsx attribute join.

Ingests the current KAIS road-name building master shapefile released
2026/03/01 and joins it against the LSMD cadastral SHP (PNU) and the
KAIS building-register XLSX (총괄표제부) so each building node carries
its parcel ID, building registry attributes, and footprint geometry.

Pipeline:
  1. Read TL_SGCO_RNADR_MST polygons (EPSG:5179, EUC-KR).
  2. Filter by SIG_CD (e.g., 30200 = Yuseong-gu).
  3. Transform centroid + boundary to WGS84.
  4. Optional: spatial-join against LSMD parcel polygons (with PNU) to assign
     each building its parcel PNU.
  5. Optional: attribute-join against 02. 총괄표제부 xlsx using PNU to enrich
     building nodes with floors, height, usage, area, approval-date, etc.

Output building dict (downstream-compatible with the legacy schema):
    uid, name, building_type, floors, underground_floors, height, elevation,
    width, depth, longitude, latitude, altitude, heading, pitch, roll, color,
    lod_level, importance, boundary,
    nf_id, address, road_address, usage_code, usage_name, structure_type,
    building_area, gross_floor_area, household_count, ho_count,
    approval_date, admin_dong, pnu
"""
from __future__ import annotations

import json
import math
import os
from typing import Iterable

import openpyxl
import shapefile
from pyproj import Transformer

# ──────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────

# All new KAIS road-name datasets ship as EPSG:5179 (UTM-K, GRS80).
_SHP_SRC_CRS = "EPSG:5179"
_TRANSFORMER = Transformer.from_crs(_SHP_SRC_CRS, "EPSG:4326", always_xy=True)

# Korean building usage code -> internal building_type bucket
USAGE_TO_TYPE = {
    "01000": "residential", "02000": "residential", "03000": "commercial",
    "04000": "commercial",  "05000": "commercial",  "06000": "religious",
    "07000": "commercial",  "08000": "commercial",  "09000": "medical",
    "10000": "education",   "11000": "commercial",  "12000": "commercial",
    "13000": "commercial",  "14000": "office",      "15000": "hotel",
    "16000": "commercial",  "17000": "industrial",  "18000": "industrial",
    "19000": "industrial",  "20000": "industrial",  "21000": "industrial",
    "22000": "industrial",  "23000": "office",      "24000": "office",
    "25000": "commercial",  "26000": "commercial",  "27000": "commercial",
    "28000": "commercial",  "29000": "commercial",
}


# ──────────────────────────────────────────────────────────────────────────
# Safe converters
# ──────────────────────────────────────────────────────────────────────────

def _safe_float(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _safe_int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _safe_str(v, default=""):
    if v is None:
        return default
    return str(v).strip()


# ──────────────────────────────────────────────────────────────────────────
# Geometry helpers
# ──────────────────────────────────────────────────────────────────────────

def _centroid_5179(points):
    """Compute centroid in source CRS (EPSG:5179), return WGS84 (lon, lat)."""
    if not points:
        return None, None
    sx = sum(p[0] for p in points)
    sy = sum(p[1] for p in points)
    n = len(points)
    cx, cy = sx / n, sy / n
    lon, lat = _TRANSFORMER.transform(cx, cy)
    return round(lon, 7), round(lat, 7)


def _polygon_to_wgs84(shape) -> list:
    """Outer ring (parts[0]:parts[1]) → [[lon, lat], ...]"""
    pts = shape.points
    parts = list(shape.parts) if shape.parts else [0]
    end = parts[1] if len(parts) > 1 else len(pts)
    out = []
    for x, y in pts[parts[0]:end]:
        lon, lat = _TRANSFORMER.transform(x, y)
        out.append([round(lon, 7), round(lat, 7)])
    return out


def _polygon_area_5179(points):
    """Shoelace area in projected CRS (m²)."""
    if len(points) < 3:
        return 0.0
    a = 0.0
    for i in range(len(points) - 1):
        a += points[i][0] * points[i + 1][1] - points[i + 1][0] * points[i][1]
    return abs(a) * 0.5


def _oriented_bbox_5179(points):
    """Approximate oriented bbox width, depth, heading (m, m, °).

    Uses the standard rotating-calipers idea but very small: scan a handful
    of angles and pick the one minimizing AABB area on rotated points.
    """
    if len(points) < 3:
        return 1.0, 1.0, 0.0
    best_w, best_d, best_angle = 1.0, 1.0, 0.0
    best_area = float("inf")
    for deg in range(0, 180, 5):
        rad = math.radians(deg)
        c, s = math.cos(rad), math.sin(rad)
        xs = [p[0] * c + p[1] * s for p in points]
        ys = [-p[0] * s + p[1] * c for p in points]
        w = max(xs) - min(xs)
        d = max(ys) - min(ys)
        a = w * d
        if a < best_area:
            best_area = a
            best_w = max(w, 1.0)
            best_d = max(d, 1.0)
            best_angle = deg
    # heading: long axis vs north (CW from north). long axis is whichever of w/d is bigger.
    if best_w >= best_d:
        return round(best_w, 1), round(best_d, 1), round(best_angle, 1)
    else:
        # if depth is the longer one, rotate by 90°
        return round(best_d, 1), round(best_w, 1), round((best_angle + 90) % 180, 1)


# ──────────────────────────────────────────────────────────────────────────
# 총괄표제부 xlsx loader
# ──────────────────────────────────────────────────────────────────────────

def load_summary_titles(xlsx_path: str) -> dict[str, dict]:
    """Load 02. 총괄표제부 xlsx → dict keyed by PNU.

    PNU (19 digits) = 시군구코드(5) + 법정동코드(5) + 대지구분(1) + 번(4) + 지(4)

    Returns: dict[pnu] -> dict with selected attributes.
    """
    if not os.path.exists(xlsx_path):
        print(f"  [총괄표제부] xlsx not found: {xlsx_path} — skipping")
        return {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(header)}

    def g(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    result: dict[str, dict] = {}
    # Per-BJD-code legal-dong name lookup, derived from "대지위치" tokens.
    # The 3rd whitespace-separated token of "대지위치" is the legal-dong
    # name (e.g., "대전광역시 유성구 원내동 1번지" → "원내동").
    bjd_to_dong_name: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sig = _safe_str(g(row, "시군구코드"))
        leg = _safe_str(g(row, "법정동코드"))
        dvs = _safe_str(g(row, "대지구분코드"))
        bn  = _safe_str(g(row, "번"))
        bu  = _safe_str(g(row, "지"))
        if not (sig and leg and bn):
            continue
        # zero-pad
        sig = sig.zfill(5)
        leg = leg.zfill(5)
        # 대지구분코드 normalisation: xlsx encodes 0=일반/1=산/2=기타,
        # while LSMD PNU's position-11 mountain flag uses 1=일반/2=산.
        # +1 shift aligns the two conventions; clamp to [1,9] for safety.
        try:
            dvs_int = int(dvs) if dvs else 0
        except ValueError:
            dvs_int = 0
        dvs = str(min(dvs_int + 1, 9))
        bn  = bn.zfill(4)
        bu  = (bu or "0").zfill(4)
        pnu = sig + leg + dvs + bn + bu  # 19 chars

        # Extract legal-dong (법정동) name from 대지위치 column.
        # Format varies by region:
        #   "대전광역시 유성구 원내동 1번지"     → 4-token: 시도 / 시군구 / 법정동 / 번지
        #   "세종특별자치시 반곡동 838번지"       → 3-token: 시도(특자시) / 법정동 / 번지
        # Heuristic: pick the first token ending in "동"/"리"/"읍"/"면" (Korean
        # cadastral-unit suffix) that comes after the 시도/시군구 prefix.
        addr_jibun = _safe_str(g(row, "대지위치"))
        legal_dong_name = ""
        if addr_jibun:
            tokens = addr_jibun.split()
            for tok in tokens:
                if tok and tok[-1] in {"동", "리", "읍", "면"} and not tok.endswith("번지"):
                    legal_dong_name = tok
                    bjd_to_dong_name.setdefault(sig + leg, legal_dong_name)
                    break

        result[pnu] = {
            "pnu": pnu,
            "legal_dong_name": legal_dong_name,
            "address_jibun":   addr_jibun,
            "address_road":    _safe_str(g(row, "도로명대지위치")),
            "building_name":   _safe_str(g(row, "건물명")),
            "site_area":       _safe_float(g(row, "대지면적(㎡)")),
            "building_area":   _safe_float(g(row, "건축면적(㎡)")),
            "coverage_ratio":  _safe_float(g(row, "건폐율(%)")),
            "gross_floor_area": _safe_float(g(row, "연면적(㎡)")),
            "floor_area_ratio": _safe_float(g(row, "용적률(%)")),
            "usage_code":      _safe_str(g(row, "주용도코드")),
            "usage_name":      _safe_str(g(row, "주용도코드명")),
            "etc_usage":       _safe_str(g(row, "기타용도")),
            "household_count": _safe_int(g(row, "세대수(세대)")),
            "family_count":    _safe_int(g(row, "가구수(가구)")),
            "main_bldg_count": _safe_int(g(row, "주건축물수")),
            "sub_bldg_count":  _safe_int(g(row, "부속건축물수")),
            "sub_bldg_area":   _safe_float(g(row, "부속건축물면적(㎡)")),
            "parking_total":   _safe_int(g(row, "총주차수")),
            "approval_date":   _safe_str(g(row, "사용승인일")),
            "permit_date":     _safe_str(g(row, "허가일")),
            "energy_grade":    _safe_str(g(row, "에너지효율등급")),
            "ho_count":        _safe_int(g(row, "호수(호)")),
            "creation_date":   _safe_str(g(row, "생성일자")),
        }
    wb.close()
    print(f"  [총괄표제부] {os.path.basename(xlsx_path)}: {len(result):,} rows indexed by PNU "
          f"({len(bjd_to_dong_name)} legal-dongs)")
    # Stash BJD→dong-name map on the dict so callers can look up dong names
    # for buildings whose summary record is missing (most buildings, since
    # 총괄표제부 only covers ~7-14% of TL_SGCO buildings).
    result["__bjd_to_dong_name__"] = bjd_to_dong_name
    return result


# ──────────────────────────────────────────────────────────────────────────
# LSMD parcel index for building → PNU spatial join
# ──────────────────────────────────────────────────────────────────────────

class _ParcelIndex:
    """Grid-bucketed parcel polygons for O(1) point-in-polygon lookup.

    Stores parcels in their native CRS (5174/5186) — buildings must
    transform to the same CRS before lookup. Caller provides parcels
    pre-transformed to EPSG:5179 (UTM-K) to match the building SHP.
    """

    def __init__(self, parcels_5179: list[dict], cell_m: float = 200.0):
        self.cell = cell_m
        self.grid: dict[tuple[int, int], list[dict]] = {}
        for p in parcels_5179:
            ring = p.get("ring_5179") or []
            if len(ring) < 3:
                continue
            xs = [pt[0] for pt in ring]
            ys = [pt[1] for pt in ring]
            x0, x1 = min(xs), max(xs)
            y0, y1 = min(ys), max(ys)
            gx0, gx1 = int(x0 // cell_m), int(x1 // cell_m)
            gy0, gy1 = int(y0 // cell_m), int(y1 // cell_m)
            for gx in range(gx0, gx1 + 1):
                for gy in range(gy0, gy1 + 1):
                    self.grid.setdefault((gx, gy), []).append(p)

    @staticmethod
    def _point_in_ring(x: float, y: float, ring: list) -> bool:
        n = len(ring)
        inside = False
        j = n - 1
        for i in range(n):
            xi, yi = ring[i]
            xj, yj = ring[j]
            if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi + 1e-12) + xi):
                inside = not inside
            j = i
        return inside

    def lookup_pnu(self, x: float, y: float) -> str | None:
        gx = int(x // self.cell)
        gy = int(y // self.cell)
        bucket = self.grid.get((gx, gy), [])
        for p in bucket:
            if self._point_in_ring(x, y, p["ring_5179"]):
                return p.get("pnu")
        return None


def _build_parcel_index(parcels: Iterable[dict],
                        src_crs: str = "EPSG:5186") -> _ParcelIndex | None:
    """Project parcel rings from src_crs to EPSG:5179 then build the index."""
    if not parcels:
        return None
    tx = Transformer.from_crs(src_crs, "EPSG:5179", always_xy=True)
    enriched = []
    for p in parcels:
        ring = p.get("boundary_5174") or p.get("boundary_5186") or p.get("ring") or []
        if not ring:
            # Fall back to WGS84 if that's all we have
            wgs = p.get("boundary")
            if isinstance(wgs, str):
                try:
                    wgs = json.loads(wgs)
                except Exception:
                    wgs = []
            if not wgs:
                continue
            ring_5179 = []
            inv_tx = Transformer.from_crs("EPSG:4326", "EPSG:5179", always_xy=True)
            for lon, lat in wgs:
                x, y = inv_tx.transform(lon, lat)
                ring_5179.append((x, y))
        else:
            ring_5179 = [tx.transform(pt[0], pt[1]) for pt in ring]
        enriched.append({"pnu": p.get("pnu") or p.get("uid"), "ring_5179": ring_5179})
    return _ParcelIndex(enriched)


# ──────────────────────────────────────────────────────────────────────────
# Main loader
# ──────────────────────────────────────────────────────────────────────────

def load_buildings(shp_path: str,
                   sig_cd_filter: str | None = None,
                   summary_xlsx: str | None = None,
                   parcels: Iterable[dict] | None = None,
                   parcel_src_crs: str = "EPSG:5186") -> list[dict]:
    """Load buildings from TL_SGCO_RNADR_MST + optional 총괄표제부 enrichment.

    Args:
        shp_path:        Path to TL_SGCO_RNADR_MST.shp (without extension OK)
        sig_cd_filter:   e.g., "30200" for Yuseong-gu, "36110" for Sejong.
                         None disables filtering (use entire SHP).
        summary_xlsx:    Path to 02. 총괄표제부 *.xlsx — optional attribute
                         enrichment by PNU. Skipped if None.
        parcels:         Iterable of parcel dicts (need 'pnu' + ring), used
                         to assign each building a PNU via spatial join.
        parcel_src_crs:  CRS of parcel rings (default EPSG:5186 for LSMD).

    Returns:
        List of building dicts.
    """
    if shp_path.endswith(".shp"):
        shp_path = shp_path[:-4]

    # 1) Pre-load summary attributes if requested
    summary: dict[str, dict] = load_summary_titles(summary_xlsx) if summary_xlsx else {}
    # Pop the BJD→dong-name lookup so it doesn't get treated as a PNU row
    bjd_to_dong_name: dict[str, str] = summary.pop("__bjd_to_dong_name__", {})

    # 2) Build parcel spatial index if requested
    parcel_idx: _ParcelIndex | None = None
    if parcels:
        parcel_idx = _build_parcel_index(parcels, src_crs=parcel_src_crs)
        if parcel_idx:
            print(f"  [Parcel index] built ({sum(len(v) for v in parcel_idx.grid.values()):,} bucket entries)")

    # 3) Stream the building SHP
    sf = shapefile.Reader(shp_path, encoding="euc-kr")
    fnames = [f[0] for f in sf.fields[1:]]
    i_sig = fnames.index("SIG_CD") if "SIG_CD" in fnames else None
    i_adr = fnames.index("ADR_MNG_NO") if "ADR_MNG_NO" in fnames else None
    i_eqb = fnames.index("EQB_MAN_SN") if "EQB_MAN_SN" in fnames else None
    i_se  = fnames.index("BULD_SE_CD") if "BULD_SE_CD" in fnames else None

    out: list[dict] = []
    skipped_geom = 0
    pnu_match = 0
    summary_match = 0

    for i, sr in enumerate(sf.iterShapeRecords()):
        rec = sr.record
        shape = sr.shape

        if i_sig is not None and sig_cd_filter and rec[i_sig] != sig_cd_filter:
            continue

        pts = [(p[0], p[1]) for p in shape.points]
        if not pts:
            skipped_geom += 1
            continue
        cent_x = sum(p[0] for p in pts) / len(pts)
        cent_y = sum(p[1] for p in pts) / len(pts)
        lon, lat = _centroid_5179(pts)
        if lon is None:
            skipped_geom += 1
            continue

        adr = _safe_str(rec[i_adr]) if i_adr is not None else ""
        eqb = _safe_str(rec[i_eqb]) if i_eqb is not None else ""
        buld_se = _safe_str(rec[i_se]) if i_se is not None else "0"
        uid = adr or eqb or f"bld-{i:06d}"

        # 4) Spatial-join PNU
        pnu = ""
        if parcel_idx is not None:
            pnu_lookup = parcel_idx.lookup_pnu(cent_x, cent_y)
            if pnu_lookup:
                pnu = pnu_lookup
                pnu_match += 1

        # 5) Summary attribute attach
        s = summary.get(pnu) if pnu else None
        # 法정동 (legal-dong) name from BJD code prefix (always available
        # when PNU spatial-join hit, even if summary record is missing).
        legal_dong_name = bjd_to_dong_name.get(pnu[:10], "") if pnu else ""

        if s:
            summary_match += 1
            usage_code = s.get("usage_code", "")
            usage_name = s.get("usage_name", "")
            bdar = s.get("building_area", 0.0)
            gfa = s.get("gross_floor_area", 0.0)
            building_name = s.get("building_name", "")
            household_count = s.get("household_count", 0)
            ho_count = s.get("ho_count", 0)
            approval_date = s.get("approval_date", "")
            road_addr = s.get("address_road", "")
            jibun_addr = s.get("address_jibun", "")
            etc_usage = s.get("etc_usage", "")
            energy_grade = s.get("energy_grade", "")
            main_bldg_count = s.get("main_bldg_count", 0)
            sub_bldg_count = s.get("sub_bldg_count", 0)
            parking_total = s.get("parking_total", 0)
            # Prefer summary-record's legal_dong_name (most accurate)
            if s.get("legal_dong_name"):
                legal_dong_name = s["legal_dong_name"]
        else:
            usage_code = ""
            usage_name = ""
            bdar = 0.0
            gfa = 0.0
            building_name = ""
            household_count = 0
            ho_count = 0
            approval_date = ""
            road_addr = ""
            jibun_addr = ""
            etc_usage = ""
            energy_grade = ""
            main_bldg_count = 0
            sub_bldg_count = 0
            parking_total = 0

        # Derived: building_type
        building_type = USAGE_TO_TYPE.get(usage_code, "commercial")
        if "아파트" in etc_usage or usage_code == "02000":
            building_type = "apartment"

        # Footprint area & dimensions
        if bdar > 0:
            target_area = bdar
        else:
            target_area = _polygon_area_5179(pts)
            if target_area <= 0:
                target_area = 25.0  # fallback 5×5

        w, d, heading = _oriented_bbox_5179(pts)
        aspect = w / d if d > 0 else 1.0
        aspect = min(max(aspect, 0.2), 5.0)
        depth = round(math.sqrt(target_area / aspect), 1)
        width = round(depth * aspect, 1)
        width = max(width, 2.0)
        depth = max(depth, 2.0)

        # Height: no direct height field on TL_SGCO_RNADR_MST; fallback by floors estimate from gfa
        # If gross floor area ≈ N × building_area then floors ≈ gfa / building_area
        if gfa > 0 and bdar > 0:
            floors = max(int(round(gfa / bdar)), 1)
        else:
            floors = 1
        height = round(floors * 3.0, 2)

        # Importance score
        importance = 0
        if floors >= 15:
            importance = 3
        elif floors >= 5:
            importance = 2
        elif floors >= 2:
            importance = 1
        if building_name:
            importance += 1
        if gfa > 5000:
            importance += 1

        boundary_coords = _polygon_to_wgs84(shape)

        out.append({
            "uid": uid,
            "name": building_name or road_addr or jibun_addr or uid,
            "building_type": building_type,
            "floors": floors,
            "underground_floors": 0,  # not present in TL_SGCO_RNADR_MST
            "height": height,
            "elevation": 0.0,
            "width": width,
            "depth": depth,
            "longitude": lon,
            "latitude": lat,
            "altitude": 0,
            "heading": heading,
            "pitch": 0,
            "roll": 0,
            "color": "",
            "lod_level": 0,
            "importance": importance,
            "boundary": json.dumps(boundary_coords),
            # Korean building registry fields
            "nf_id": uid,
            "address": jibun_addr,
            "road_address": road_addr,
            "usage_code": usage_code,
            "usage_name": usage_name,
            "structure_type": "",
            "building_area": round(bdar, 2),
            "gross_floor_area": round(gfa, 2),
            "household_count": household_count,
            "ho_count": ho_count,
            "approval_date": approval_date,
            "admin_dong": legal_dong_name,
            "legal_dong_name": legal_dong_name,
            "legal_dong_code": pnu[:10] if pnu else "",
            "pnu": pnu,
            # KAIS RN_CD (road-name code) extracted from ADR_MNG_NO positions
            # [8:15] — used by the ON_STREET attribute-match rule to join this
            # building to the road of the same RN_CD value.
            "rn_cd": adr[8:15] if len(adr) >= 15 else "",
            # Extended attributes from 총괄표제부 (new)
            "energy_grade": energy_grade,
            "main_bldg_count": main_bldg_count,
            "sub_bldg_count": sub_bldg_count,
            "parking_total": parking_total,
            "etc_usage": etc_usage,
            "buld_se_cd": buld_se,
            "eqb_man_sn": eqb,
        })

        if (i + 1) % 10000 == 0:
            print(f"  [SHP] Scanned {i + 1:,} / {sf.__len__():,} records...")

    sf.close()

    print(f"  [SHP] Buildings emitted: {len(out):,}"
          + (f" (filter SIG_CD={sig_cd_filter})" if sig_cd_filter else ""))
    if skipped_geom:
        print(f"  [SHP] Skipped {skipped_geom:,} records with empty/bad geometry")
    if parcel_idx is not None:
        print(f"  [PNU spatial join] matched {pnu_match:,} / {len(out):,} "
              f"({pnu_match/max(len(out),1)*100:.1f}%)")
    if summary:
        print(f"  [총괄표제부 attribute join] matched {summary_match:,} / {len(out):,} "
              f"({summary_match/max(len(out),1)*100:.1f}%)")

    return out
